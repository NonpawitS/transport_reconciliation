"""
SPX Transport Order (TO) Excel Parser
Format: .xlsx — export จากระบบ SPX ("transport_list_*.xlsx")

Structure:
  Row 0-21 : metadata block (คู่ key/value ใน column A/B)
             TO Number, Total Orders, TO Created Time, TO Complete Time, ...
  Row 22   : blank
  Row 23   : header — No | LM Tracking Number | District | Scan Time |
                      Weight(kg) | Journey Type | Remark
  Row 24+  : data rows

หมายเหตุสำคัญ:
  ไฟล์นี้ "ไม่มี Order SN" (ต่างจาก SPX PDF ที่มีทั้ง tracking และ order_sn)
  → order_sn จะเป็น "" ทุก row
  → การ reconcile ต้องพึ่ง Tracking Number เป็นหลัก (TLD และ FC รองรับทั้งคู่)

Output schema ตรงกับ parse_spx_pdf() เพื่อให้ concat รวมกันได้:
  tracking, order_sn, pickup_time  (+ to_number, district, weight, journey_type, page)
"""
import io
from datetime import date, datetime, time

import openpyxl
import pandas as pd

# header ที่ใช้ระบุตำแหน่งตารางข้อมูล
_HEADER_FIRST_CELL = "No"
_TRACKING_HEADER_HINT = "Tracking"

# mapping: header ในไฟล์ → column name ที่เราใช้
_COLUMN_MAP = {
    "LM Tracking Number": "tracking",
    "Tracking Number":    "tracking",
    "District":           "district",
    "Scan Time":          "pickup_time",
    "Weight(kg)":         "weight",
    "Journey Type":       "journey_type",
    "Remark":             "remark",
}

OUTPUT_COLUMNS = [
    "tracking", "order_sn", "pickup_time",
    "to_number", "district", "weight", "journey_type", "page",
]


def _cell_str(v) -> str:
    """แปลงค่า cell → string สะอาด (datetime → 'YYYY-MM-DD HH:MM:SS')"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    s = str(v).strip()
    return "" if s.lower() == "none" else s


def _find_header_row(rows: list[tuple]) -> int:
    """หาแถว header ของตารางข้อมูล — return index หรือ -1 ถ้าไม่พบ"""
    for i, row in enumerate(rows):
        if not row:
            continue
        cells = [_cell_str(v) for v in row]
        if cells and cells[0] == _HEADER_FIRST_CELL and any(
            _TRACKING_HEADER_HINT in c for c in cells[1:]
        ):
            return i
    return -1


def _parse_metadata(rows: list[tuple], stop_at: int) -> dict[str, str]:
    """อ่าน metadata block (key ใน column A, value ใน column B) ก่อนถึงแถว header"""
    meta: dict[str, str] = {}
    for row in rows[:stop_at]:
        if not row or len(row) < 2:
            continue
        key = _cell_str(row[0])
        val = _cell_str(row[1])
        if key and key not in meta:
            meta[key] = val
    return meta


def parse_spx_xlsx(file_bytes: bytes) -> tuple[pd.DataFrame, str]:
    """
    Parse SPX Transport Order Excel file.
    Returns (DataFrame, to_number)

    DataFrame columns: tracking, order_sn, pickup_time,
                       to_number, district, weight, journey_type, page
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    empty = pd.DataFrame(columns=OUTPUT_COLUMNS)

    hdr_idx = _find_header_row(rows)
    if hdr_idx < 0:
        return empty, ""

    meta = _parse_metadata(rows, hdr_idx)
    to_number = meta.get("TO Number", "")

    header = [_cell_str(v) for v in rows[hdr_idx]]
    # ตำแหน่ง column ที่สนใจ
    col_pos = {
        _COLUMN_MAP[h]: i
        for i, h in enumerate(header)
        if h in _COLUMN_MAP
    }
    if "tracking" not in col_pos:
        return empty, to_number

    data = []
    for row in rows[hdr_idx + 1:]:
        if not row:
            continue
        tracking = _cell_str(row[col_pos["tracking"]]) if col_pos["tracking"] < len(row) else ""
        if not tracking:
            continue

        def _get(key: str) -> str:
            pos = col_pos.get(key)
            return _cell_str(row[pos]) if pos is not None and pos < len(row) else ""

        data.append({
            "tracking":     tracking,
            "order_sn":     "",              # SPX TO export ไม่มี Order SN
            "pickup_time":  _get("pickup_time"),
            "to_number":    to_number,
            "district":     _get("district"),
            "weight":       _get("weight"),
            "journey_type": _get("journey_type"),
            "page":         1,
        })

    if not data:
        return empty, to_number

    df = pd.DataFrame(data, columns=OUTPUT_COLUMNS)
    df = df.drop_duplicates(subset=["tracking"]).reset_index(drop=True)
    return df, to_number


def get_spx_xlsx_to_number(file_bytes: bytes) -> str:
    """อ่านเฉพาะ TO Number จาก metadata (เร็วกว่า parse ทั้งไฟล์)"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(max_row=30, values_only=True))
    finally:
        wb.close()
    for row in rows:
        if row and len(row) >= 2 and _cell_str(row[0]) == "TO Number":
            return _cell_str(row[1])
    return ""


def validate_spx_xlsx_df(df: pd.DataFrame) -> tuple[bool, str]:
    """Validate parsed SPX Excel DataFrame. Returns (is_valid, error_message)"""
    if df.empty:
        return False, "ไม่พบข้อมูลในไฟล์ Excel — ตรวจสอบว่าเป็น SPX Transport Order export"
    if "tracking" not in df.columns:
        return False, "ไม่พบ Column 'LM Tracking Number'"
    if df["tracking"].astype(str).str.strip().eq("").all():
        return False, "Tracking Number ว่างทุก Row — ตรวจสอบ Format ของไฟล์"
    return True, ""
